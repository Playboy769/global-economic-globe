"""
EMA Bias 加倉模型 — 全流程一站式執行 (v3 / End-to-End)

從 Yahoo Finance 抓資料 → 套用 EMA Bias 評分模型 → 產生 3 個矩陣
→ Version A / B 效率前緣最佳化 → 輸出 CSV + 對照圖

不再需要 Excel/VBA 即可跑完整流程。
若想沿用 VBA 匯出的 CSV，請執行: python portfolio_optimizer.py --use-csv

Version A (baseline):
  - Σ 從 model returns (raw × Position) 估計
  - 風險指標含全部交易日 → Sharpe 通常虛高

Version B (improved):
  - Σ 從 raw returns 估計 + Ledoit-Wolf shrinkage
  - Position 用 5 日 SMA 平滑
  - 風險指標只算「組合至少有部分曝險」的日子
  - 預期 Sharpe 0.5-2.5、MaxDD -10%~-25% 是合理區間

輸出: portfolio_*.csv (3 個矩陣)、frontier_A.csv、frontier_B.csv、frontier_compare.png
"""

import os
import sys
import time
import argparse
from typing import Optional

import requests
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.covariance import LedoitWolf
import cvxpy as cp


# ============================================================
# 設定區 (要改參數改這裡)
# ============================================================

# 股票池來源 — 預設讀 Efficiency.xlsm 的 Portfolio 分頁 A 欄
# 找不到檔案時 fallback 用下方 TICKERS_FALLBACK
EXCEL_PATH = "Efficiency.xlsm"      # 相對於本 .py 同目錄
EXCEL_SHEET = "Portfolio"
EXCEL_TICKER_COL = "A"              # ticker 所在欄
EXCEL_HEADER_ROW = 1                # 第 1 列是標題 "Ticker"，從第 2 列開始讀

TICKERS_FALLBACK = [
    "2330.TW", "2454.TW", "2317.TW", "2308.TW", "2382.TW",
    "3008.TW", "2412.TW", "1301.TW", "1303.TW", "2303.TW",
    "2891.TW", "2882.TW", "0050.TW", "0056.TW", "00878.TW",
]

LOOKBACK_YEARS = 1     # Yahoo range=1y 約 252 個交易日
INDEX_TWII = "^TWII"
INDEX_OTC = "^TWOII"

# 最佳化參數
TRADING_DAYS = 252
TRAIN_RATIO = 0.7
N_FRONTIER_POINTS = 25
POSITION_SMA_WINDOW = 5
OVERFIT_GAP_THRESHOLD = 0.05
MAX_WEIGHT: Optional[float] = None   # 例如 0.15 → 單檔上限 15%。None 不設限

# Yahoo Finance
YAHOO_URL = "https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?range={years}y&interval=1d"
YAHOO_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}
FETCH_RETRY = 3
FETCH_SLEEP = 1.5    # 每檔間隔避免 rate limit

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ============================================================
# Stage 0 — 從 Excel Portfolio 分頁讀股票清單
# ============================================================

def load_tickers_from_excel(path: str = EXCEL_PATH,
                            sheet: str = EXCEL_SHEET,
                            col: str = EXCEL_TICKER_COL,
                            header_row: int = EXCEL_HEADER_ROW) -> list:
    """從 Excel 讀 Portfolio 分頁的 ticker 清單 (跳過標題列、去空白/重複)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] 未安裝 openpyxl，請執行: pip install openpyxl")
        print("       或在命令列加 --tickers 2330.TW,2454.TW 直接指定")
        sys.exit(1)

    abs_path = path if os.path.isabs(path) else os.path.join(OUTPUT_DIR, path)
    if not os.path.isfile(abs_path):
        print(f"[WARN] 找不到 Excel 檔: {abs_path}")
        print(f"       改用 fallback 清單 ({len(TICKERS_FALLBACK)} 檔)")
        return list(TICKERS_FALLBACK)

    print(f"\n=== Stage 0: 從 Excel 讀 ticker ===")
    print(f"  檔案: {abs_path}")
    print(f"  分頁: {sheet} / 欄: {col} (跳過第 {header_row} 列標題)")

    wb = load_workbook(abs_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        print(f"[ERROR] Excel 中找不到 '{sheet}' 分頁，現有分頁: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet]

    tickers, seen = [], set()
    for row in ws.iter_rows(min_row=header_row + 1, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if not s or s.upper() == "TICKER":
            continue
        if s in seen:
            continue
        seen.add(s)
        tickers.append(s)
    wb.close()

    if not tickers:
        print(f"[WARN] Portfolio 分頁 {col} 欄沒有有效 ticker，改用 fallback 清單")
        return list(TICKERS_FALLBACK)

    print(f"  讀到 {len(tickers)} 檔: {', '.join(tickers[:5])}" +
          (f" ... +{len(tickers)-5} 檔" if len(tickers) > 5 else ""))
    return tickers


# ============================================================
# Stage 1 — Yahoo Finance 資料抓取
# ============================================================

def fetch_ohlcv(ticker: str, years: int = LOOKBACK_YEARS) -> Optional[pd.DataFrame]:
    """從 Yahoo v8 chart API 抓 daily OHLCV，回傳 DataFrame (index = Date)"""
    url = YAHOO_URL.format(ticker=ticker, years=years)
    last_err = ""
    for attempt in range(FETCH_RETRY):
        try:
            r = requests.get(url, headers=YAHOO_HEADERS, timeout=20)
            if r.status_code == 200:
                break
            last_err = f"HTTP {r.status_code}"
        except Exception as e:
            last_err = str(e)
        time.sleep(2)
    else:
        print(f"    fetch fail ({last_err})")
        return None

    try:
        result = r.json()['chart']['result'][0]
        ts = result['timestamp']
        ind = result['indicators']['quote'][0]
        df = pd.DataFrame({
            'Date': pd.to_datetime(ts, unit='s').normalize(),
            'Open':   ind.get('open',   []),
            'High':   ind.get('high',   []),
            'Low':    ind.get('low',    []),
            'Close':  ind.get('close',  []),
            'Volume': ind.get('volume', []),
        }).set_index('Date')
        df = df.dropna(subset=['Close'])
        df = df[df['Close'] > 0]
        return df
    except Exception as e:
        print(f"    JSON parse fail: {e}")
        return None


# ============================================================
# Stage 2 — EMA Bias 評分模型 (Python 版，邏輯對齊 VBA)
# ============================================================

def is_taiwan(ticker: str) -> bool:
    t = ticker.upper()
    return t.endswith('.TW') or t.endswith('.TWO')


def order_score(close, e6, e13, e30, s60) -> np.ndarray:
    """排列分 (4 階梯): 6 + 6 + 6 + 7 = 25 滿分"""
    s = np.where(close > e6, 6, 0)
    s = s + np.where(e6 > e13, 6, 0)
    s = s + np.where(e13 > e30, 6, 0)
    s = s + np.where(e30 > s60, 7, 0)
    return s.astype(float)


def deviation_score(bias: np.ndarray) -> np.ndarray:
    """SMA60 乖離分 (0~20)，越遠越好 (僅獎勵正向乖離)"""
    out = np.full(len(bias), np.nan)
    for i, b in enumerate(bias):
        if pd.isna(b):
            continue
        if   b > 20:   out[i] = 20
        elif b > 15:   out[i] = 17
        elif b > 10:   out[i] = 14
        elif b > 5:    out[i] = 11
        elif b > 0:    out[i] = 7
        else:          out[i] = 3     # 在 SMA60 下方時僅給底分
    return out


def pullback_score(pullback: np.ndarray) -> np.ndarray:
    """30 日高點回撤分 (0~20)，pullback 為 <=0 的百分比 (越接近 0 越好)"""
    out = np.full(len(pullback), np.nan)
    for i, p in enumerate(pullback):
        if pd.isna(p):
            continue
        pa = -p   # 轉正值 (0 = 創新高，正數 = 回撤幅度)
        if   pa <= 2:   out[i] = 20
        elif pa <= 5:   out[i] = 16
        elif pa <= 10:  out[i] = 12
        elif pa <= 15:  out[i] = 8
        elif pa <= 25:  out[i] = 4
        else:           out[i] = 1
    return out


def vol_score(ratio: np.ndarray) -> np.ndarray:
    """量能分 (0~35)，今日量 / 20MA"""
    out = np.full(len(ratio), np.nan)
    for i, r in enumerate(ratio):
        if pd.isna(r): continue
        if   r > 4:    out[i] = 10
        elif r > 2.5:  out[i] = 22
        elif r > 1.5:  out[i] = 35
        elif r > 1:    out[i] = 26
        elif r > 0.7:  out[i] = 15
        else:          out[i] = 5
    return out


def build_index_features(idx: pd.DataFrame) -> pd.DataFrame:
    """指數: 加 EMA20 + 連續下降天數"""
    df = idx.copy()
    df['EMA20'] = df['Close'].ewm(span=20, adjust=False).mean()
    df['DeclineFlag'] = (df['EMA20'] < df['EMA20'].shift(1)).astype(int)
    days, cnt = [], 0
    for f in df['DeclineFlag']:
        cnt = cnt + 1 if f == 1 else 0
        days.append(cnt)
    df['DeclineDays'] = days
    return df


def build_model(data: pd.DataFrame, twii: pd.DataFrame, otc: pd.DataFrame,
                ticker: str) -> pd.DataFrame:
    """對單一標的計算所有指標與分數，回傳完整 model DataFrame"""
    df = data.copy()
    df['EMA6']  = df['Close'].ewm(span=6, adjust=False).mean()
    df['EMA13'] = df['Close'].ewm(span=13, adjust=False).mean()
    df['EMA30'] = df['Close'].ewm(span=30, adjust=False).mean()
    df['SMA60'] = df['Close'].rolling(60).mean()
    df['VolSMA20'] = df['Volume'].rolling(20).mean()
    df['Bias_SMA60'] = (df['Close'] - df['SMA60']) / df['SMA60'] * 100

    # 近 30 個交易日的最高價 + 從高點回撤幅度 (Pullback <= 0%)
    df['High30'] = df['High'].rolling(30).max()
    df['Pullback'] = (df['Close'] - df['High30']) / df['High30'] * 100

    df['OrderScore'] = order_score(
        df['Close'].values, df['EMA6'].values,
        df['EMA13'].values, df['EMA30'].values, df['SMA60'].values
    )
    df.loc[df['SMA60'].isna(), 'OrderScore'] = np.nan
    df['DeviationScore'] = deviation_score(df['Bias_SMA60'].values)
    df['PullbackScore'] = pullback_score(df['Pullback'].values)
    df['VolRatio'] = df['Volume'] / df['VolSMA20']
    df['VolScore'] = vol_score(df['VolRatio'].values)

    if is_taiwan(ticker):
        twii_dd = twii['DeclineDays'].reindex(df.index).ffill().fillna(0)
        otc_dd  = otc['DeclineDays'].reindex(df.index).ffill().fillna(0)
        df['TWII_DD'] = twii_dd.values
        df['OTC_DD']  = otc_dd.values
        df['IndexPenalty'] = df['TWII_DD'] * 0.75 + df['OTC_DD'] * 0.5
    else:
        df['TWII_DD'] = 0
        df['OTC_DD']  = 0
        df['IndexPenalty'] = 0

    df['Ret5D'] = (df['Close'] - df['Close'].shift(5)) / df['Close'].shift(5) * 100
    df['PanicFlag'] = ((df['IndexPenalty'] > 60) & (df['Ret5D'] < -15)).astype(int)

    # TotalScore: PanicFlag=1 → VolScore × 2、跳過 IndexPenalty
    # 滿分: 排列 25 + 乖離 20 + 回撤 20 + 量能 35 = 100
    score = np.where(
        df['PanicFlag'].values == 1,
        df['OrderScore'].values + df['DeviationScore'].values
            + df['PullbackScore'].values + df['VolScore'].values * 2,
        df['OrderScore'].values + df['DeviationScore'].values
            + df['PullbackScore'].values + df['VolScore'].values - df['IndexPenalty'].values
    )
    score = np.clip(score, 0, None)
    score[df[['OrderScore', 'DeviationScore', 'PullbackScore', 'VolScore']].isna().any(axis=1).values] = np.nan
    df['TotalScore'] = score

    df['Position'] = ((df['TotalScore'] - 40) / 60 * 100).clip(0, 100)
    return df


# ============================================================
# Stage 3 — 組合 3 個矩陣
# ============================================================

def build_portfolio_matrices(tickers):
    """跑全部 ticker，組成 (model_ret, raw_ret, positions) 三個矩陣"""
    print("\n=== Stage 1+2: 抓資料 + 跑模型 ===")
    print("  抓大盤指數...")
    twii_raw = fetch_ohlcv(INDEX_TWII)
    otc_raw  = fetch_ohlcv(INDEX_OTC)
    if twii_raw is None or otc_raw is None:
        print("[ERROR] 大盤指數抓取失敗，無法計算 IndexPenalty。請檢查網路。")
        sys.exit(1)
    twii = build_index_features(twii_raw)
    otc  = build_index_features(otc_raw)
    print(f"    TWII: {len(twii)} 筆, OTC: {len(otc)} 筆")

    raw_ret_d, model_ret_d, pos_d = {}, {}, {}
    for i, t in enumerate(tickers, 1):
        print(f"  [{i:2}/{len(tickers)}] {t:<14} ", end='', flush=True)
        data = fetch_ohlcv(t)
        time.sleep(FETCH_SLEEP)
        if data is None:
            print("SKIP (抓取失敗)")
            continue
        if len(data) < 70:
            print(f"SKIP (僅 {len(data)} 筆 < 70)")
            continue
        model = build_model(data, twii, otc, t)
        raw   = model['Close'].pct_change().fillna(0)
        prev_pos = model['Position'].shift(1).fillna(0) / 100.0
        raw_ret_d[t]   = raw
        model_ret_d[t] = raw * prev_pos
        pos_d[t]       = model['Position'].fillna(0)
        print(f"OK ({len(model)} 筆)")

    if not raw_ret_d:
        print("[ERROR] 所有股票都失敗，無法繼續")
        sys.exit(1)

    raw_ret   = pd.DataFrame(raw_ret_d).sort_index().fillna(0)
    model_ret = pd.DataFrame(model_ret_d).sort_index().fillna(0)
    positions = pd.DataFrame(pos_d).sort_index().fillna(0)
    return model_ret, raw_ret, positions


# ============================================================
# Stage 4 — CSV 模式 (fallback for --use-csv)
# ============================================================

def find_csv(name: str) -> Optional[str]:
    candidates = [
        os.path.join(OUTPUT_DIR, name),
        os.path.join(os.getcwd(), name),
        os.path.expanduser(f'~\\Desktop\\{name}'),
        os.path.join(os.environ.get('TEMP', ''), name),
        os.path.expanduser(f'~\\OneDrive\\桌面\\{name}'),
    ]
    for p in candidates:
        if p and os.path.isfile(p):
            return p
    return None


def load_csv(name: str, required: bool = True) -> Optional[pd.DataFrame]:
    p = find_csv(name)
    if p is None:
        if required:
            print(f"[ERROR] 找不到 {name}")
            print("解法: 不要加 --use-csv (預設用 Yahoo 直抓)")
            sys.exit(1)
        return None
    df = pd.read_csv(p, parse_dates=['Date']).set_index('Date').sort_index()
    df = df.apply(pd.to_numeric, errors='coerce').dropna(axis=1, how='all')
    nan_count = int(df.isna().sum().sum())
    if nan_count > 0:
        print(f"  Loaded {name}: {df.shape}  (補 0 處理 {nan_count} 個 NaN)")
        df = df.fillna(0)
    else:
        print(f"  Loaded {name}: {df.shape}")
    return df


# ============================================================
# Stage 5 — 最佳化共用工具
# ============================================================

def walkforward_split(df: pd.DataFrame, ratio: float = TRAIN_RATIO):
    cutoff = int(len(df) * ratio)
    return df.iloc[:cutoff], df.iloc[cutoff:]


def estimate_mu_sigma(returns: pd.DataFrame, shrink: bool = True):
    mu = returns.mean().values
    if shrink:
        lw = LedoitWolf().fit(returns.values)
        sigma = lw.covariance_
        print(f"  Ledoit-Wolf shrinkage: {lw.shrinkage_:.4f}")
    else:
        sigma = returns.cov().values
    return mu, sigma


def max_drawdown(rets) -> float:
    nav = (1 + pd.Series(rets)).cumprod()
    peak = nav.cummax()
    return float(((nav - peak) / peak).min())


def solve_min_variance(sigma, mu, target_ret, max_w: Optional[float] = MAX_WEIGHT):
    n = len(mu)
    w = cp.Variable(n)
    constraints = [cp.sum(w) == 1, w >= 0, mu @ w >= target_ret]
    if max_w is not None:
        constraints.append(w <= max_w)
    prob = cp.Problem(cp.Minimize(cp.quad_form(w, cp.psd_wrap(sigma))), constraints)
    try:
        prob.solve(solver=cp.SCS, verbose=False)
    except Exception:
        return None
    if w.value is None or prob.status not in ('optimal', 'optimal_inaccurate'):
        return None
    w_clean = np.maximum(w.value, 0)
    s = w_clean.sum()
    return w_clean / s if s > 0 else None


def compute_metrics(rets, mask=None) -> dict:
    active = rets[mask] if mask is not None else rets
    if len(active) < 2:
        return dict(mean=0, std=0, sharpe=0, maxdd=0)
    mean = active.mean() * TRADING_DAYS
    std  = active.std(ddof=1) * np.sqrt(TRADING_DAYS)
    sharpe = mean / std if std > 0 else 0
    maxdd = max_drawdown(rets)
    return dict(mean=mean, std=std, sharpe=sharpe, maxdd=maxdd)


# ============================================================
# Stage 6 — Version A / B 最佳化
# ============================================================

def version_a(model_ret: pd.DataFrame) -> pd.DataFrame:
    print("\n" + "=" * 60)
    print("Version A — model returns (raw × Position[t-1])")
    print("=" * 60)
    train, test = walkforward_split(model_ret)
    print(f"  Train: {len(train)} rows  Test: {len(test)} rows")
    mu, sigma = estimate_mu_sigma(train, shrink=True)

    targets = np.linspace(mu.min(), mu.max(), N_FRONTIER_POINTS)
    rows = []
    for t in targets:
        w = solve_min_variance(sigma, mu, t)
        if w is None:
            continue
        is_ret = train.values @ w
        oos_ret = test.values @ w
        is_m, oos_m = compute_metrics(is_ret), compute_metrics(oos_ret)
        row = {
            'version': 'A',
            'target_ret_annual': t * TRADING_DAYS,
            'is_ret_annual':  is_m['mean'],  'oos_ret_annual':  oos_m['mean'],
            'is_std_annual':  is_m['std'],   'oos_std_annual':  oos_m['std'],
            'is_maxdd':       is_m['maxdd'], 'oos_maxdd':       oos_m['maxdd'],
            'is_sharpe':      is_m['sharpe'],'oos_sharpe':      oos_m['sharpe'],
        }
        for col, wi in zip(model_ret.columns, w):
            row[f'w_{col}'] = float(wi)
        rows.append(row)
    return pd.DataFrame(rows)


def version_b(raw_ret: pd.DataFrame, positions: pd.DataFrame) -> pd.DataFrame:
    print("\n" + "=" * 60)
    print("Version B — raw Σ + SMA5(Position) + active-day metrics")
    print("=" * 60)

    common_cols = [c for c in raw_ret.columns if c in positions.columns]
    common_idx  = raw_ret.index.intersection(positions.index)
    raw_ret   = raw_ret.loc[common_idx, common_cols]
    positions = positions.loc[common_idx, common_cols]

    pos_smooth = positions.rolling(POSITION_SMA_WINDOW, min_periods=1).mean()
    print(f"  Position SMA{POSITION_SMA_WINDOW} 平滑後")
    print(f"  原始 0-Position 日比例: {(positions == 0).all(axis=1).mean():.2%}")
    print(f"  平滑 0-Position 日比例: {(pos_smooth == 0).all(axis=1).mean():.2%}")

    pos_lag = pos_smooth.shift(1).fillna(0) / 100.0
    eff_ret = (raw_ret * pos_lag).dropna(how='any')

    train_eff, test_eff = walkforward_split(eff_ret)
    train_raw = raw_ret.loc[train_eff.index]
    print(f"  Train: {len(train_eff)} rows  Test: {len(test_eff)} rows")

    mu_raw, sigma_raw = estimate_mu_sigma(train_raw, shrink=True)
    mu_eff = train_eff.mean().values

    targets = np.linspace(mu_eff.min(), mu_eff.max(), N_FRONTIER_POINTS)
    rows = []
    for t in targets:
        w = solve_min_variance(sigma_raw, mu_eff, t)
        if w is None:
            continue
        is_ret  = train_eff.values @ w
        oos_ret = test_eff.values @ w
        is_pos  = pos_lag.loc[train_eff.index].values @ w
        oos_pos = pos_lag.loc[test_eff.index].values @ w
        is_active  = is_pos  > 0.001
        oos_active = oos_pos > 0.001
        is_m  = compute_metrics(is_ret,  mask=is_active)
        oos_m = compute_metrics(oos_ret, mask=oos_active)

        row = {
            'version': 'B',
            'target_ret_annual': t * TRADING_DAYS,
            'is_ret_annual':  is_m['mean'],  'oos_ret_annual':  oos_m['mean'],
            'is_std_annual':  is_m['std'],   'oos_std_annual':  oos_m['std'],
            'is_maxdd':       is_m['maxdd'], 'oos_maxdd':       oos_m['maxdd'],
            'is_sharpe':      is_m['sharpe'],'oos_sharpe':      oos_m['sharpe'],
            'is_active_ratio':  is_active.mean(),
            'oos_active_ratio': oos_active.mean(),
        }
        for col, wi in zip(raw_ret.columns, w):
            row[f'w_{col}'] = float(wi)
        rows.append(row)
    return pd.DataFrame(rows)


# ============================================================
# Stage 7 — 對照圖 / 報告
# ============================================================

def plot_compare(df_a, df_b, output_path):
    fig, axes = plt.subplots(2, 2, figsize=(14, 11))

    ax = axes[0, 0]
    ax.plot(df_a['is_std_annual']*100, df_a['is_ret_annual']*100, 'o-', label='Version A', linewidth=2)
    ax.plot(df_b['is_std_annual']*100, df_b['is_ret_annual']*100, 's-', label='Version B', linewidth=2)
    ax.set(xlabel='Annualized Std (%)', ylabel='Annualized Return (%)', title='Efficient Frontier (In-sample)')
    ax.legend(); ax.grid(alpha=0.3)

    ax = axes[0, 1]
    ax.plot(df_a['oos_std_annual']*100, df_a['oos_ret_annual']*100, 'o-', label='Version A', linewidth=2)
    ax.plot(df_b['oos_std_annual']*100, df_b['oos_ret_annual']*100, 's-', label='Version B', linewidth=2)
    ax.set(xlabel='Annualized Std (%)', ylabel='Annualized Return (%)', title='Efficient Frontier (Out-of-sample)')
    ax.legend(); ax.grid(alpha=0.3)

    ax = axes[1, 0]
    ax.plot(-df_a['oos_maxdd']*100, df_a['oos_ret_annual']*100, 'o-', label='Version A', linewidth=2)
    ax.plot(-df_b['oos_maxdd']*100, df_b['oos_ret_annual']*100, 's-', label='Version B', linewidth=2)
    ax.set(xlabel='Max Drawdown (%)', ylabel='Annualized Return (%)', title='DD vs Return (OOS)')
    ax.legend(); ax.grid(alpha=0.3)

    ax = axes[1, 1]
    x = np.arange(len(df_a))
    ax.plot(x, df_a['oos_sharpe'], 'o-', label='Version A (suspect)', linewidth=2)
    ax.plot(x[:len(df_b)], df_b['oos_sharpe'].values, 's-', label='Version B (realistic)', linewidth=2)
    ax.axhline(2, color='gray', linestyle=':', label='Sharpe=2 (very good)')
    ax.axhline(1, color='gray', linestyle=':', alpha=0.5, label='Sharpe=1 (good)')
    ax.set(xlabel='Frontier point index', ylabel='OOS Sharpe', title='Sharpe Reality Check')
    ax.legend(); ax.grid(alpha=0.3)

    plt.tight_layout()
    plt.savefig(output_path, dpi=120)
    print(f"\nSaved: {output_path}")


def best_summary(df: pd.DataFrame, label: str):
    if len(df) == 0 or 'oos_sharpe' not in df.columns:
        return
    feas = df[df['oos_sharpe'].notna()]
    if len(feas) == 0:
        return
    best = feas.loc[feas['oos_sharpe'].idxmax()]
    print(f"\n--- {label} OOS-best Sharpe ---")
    print(f"  Sharpe = {best['oos_sharpe']:.3f}  Std = {best['oos_std_annual']:.4f}  MaxDD = {best['oos_maxdd']:.4f}")
    print(f"  Annual Ret = {best['oos_ret_annual']:.4f}")
    weight_cols = [c for c in df.columns if c.startswith('w_')]
    weights = sorted([(c[2:], best[c]) for c in weight_cols if best[c] > 0.005],
                     key=lambda x: -x[1])
    print("  Top weights (>0.5%):")
    for tk, w in weights[:10]:
        print(f"    {tk:<14} {w*100:6.2f}%")


def report_gap(df: pd.DataFrame, label: str):
    if len(df) == 0:
        return
    gap_std = (df['is_std_annual'] - df['oos_std_annual']).abs().mean()
    gap_ret = (df['is_ret_annual'] - df['oos_ret_annual']).abs().mean()
    flag = "  ⚠ OVERFIT" if gap_std > OVERFIT_GAP_THRESHOLD else "  ✓"
    print(f"  {label}: |IS-OOS| std gap = {gap_std:.4f}, return gap = {gap_ret:.4f}{flag}")


# ============================================================
# Main 編排
# ============================================================

def parse_args():
    p = argparse.ArgumentParser(description="EMA Bias 全流程一站式執行")
    p.add_argument('--use-csv', action='store_true',
                   help='不從 Yahoo 抓資料，改用本地 portfolio_*.csv (VBA 匯出)')
    p.add_argument('--no-save-matrices', action='store_true',
                   help='不要存 3 個矩陣 CSV (只算最佳化結果)')
    p.add_argument('--excel', default=EXCEL_PATH,
                   help=f'Excel 檔路徑 (預設: {EXCEL_PATH})')
    p.add_argument('--sheet', default=EXCEL_SHEET,
                   help=f'分頁名稱 (預設: {EXCEL_SHEET})')
    p.add_argument('--tickers', default=None,
                   help='直接指定 ticker (逗號分隔)，例如 --tickers 2330.TW,2454.TW，會覆寫 Excel 讀取')
    return p.parse_args()


def main():
    args = parse_args()
    print("=" * 60)
    print(" EMA Bias 加倉模型 — 全流程一站式執行 (v3)")
    print(f" 模式: {'CSV (VBA 匯出)' if args.use_csv else 'Yahoo Finance 直抓'}")
    print(f" 單檔權重上限: {f'{MAX_WEIGHT*100:.0f}%' if MAX_WEIGHT else '無'}")
    print("=" * 60)

    if args.use_csv:
        print("\n=== Stage 1+2+3: 載入本地 CSV ===")
        model_ret = load_csv('portfolio_returns.csv')
        raw_ret   = load_csv('portfolio_raw_returns.csv')
        positions = load_csv('portfolio_positions.csv')
    else:
        # 取得 ticker 清單: --tickers 優先 → Excel → fallback
        if args.tickers:
            tickers = [t.strip() for t in args.tickers.split(',') if t.strip()]
            print(f"\n=== Stage 0: 命令列指定 {len(tickers)} 檔 ticker ===")
            print(f"  {', '.join(tickers)}")
        else:
            tickers = load_tickers_from_excel(path=args.excel, sheet=args.sheet)
        if not tickers:
            print("[ERROR] 沒有可用的 ticker，結束")
            sys.exit(1)
        model_ret, raw_ret, positions = build_portfolio_matrices(tickers)
        if not args.no_save_matrices:
            print("\n  存 3 個矩陣 CSV...")
            model_ret.to_csv(os.path.join(OUTPUT_DIR, 'portfolio_returns.csv'),
                             float_format='%.8f', date_format='%Y-%m-%d')
            raw_ret.to_csv(os.path.join(OUTPUT_DIR, 'portfolio_raw_returns.csv'),
                           float_format='%.8f', date_format='%Y-%m-%d')
            positions.to_csv(os.path.join(OUTPUT_DIR, 'portfolio_positions.csv'),
                             float_format='%.4f', date_format='%Y-%m-%d')
            print(f"  Saved: portfolio_returns / raw_returns / positions .csv ({model_ret.shape})")

    df_a = version_a(model_ret)
    df_b = version_b(raw_ret, positions)

    df_a.to_csv(os.path.join(OUTPUT_DIR, 'frontier_A.csv'), index=False)
    df_b.to_csv(os.path.join(OUTPUT_DIR, 'frontier_B.csv'), index=False)
    print(f"\nSaved: frontier_A.csv, frontier_B.csv")

    plot_compare(df_a, df_b, os.path.join(OUTPUT_DIR, 'frontier_compare.png'))

    print("\n=== 過配適 (overfit) gap 對照 ===")
    report_gap(df_a, "Version A")
    report_gap(df_b, "Version B")

    best_summary(df_a, "Version A")
    best_summary(df_b, "Version B")

    print("\n=== 判讀提示 ===")
    print("  Version A Sharpe 通常虛高 (>5), Version B 應落在 0.5-2.5 才合理")
    print("  Version A MaxDD 偏低 (<5%), Version B 應落在 -10% ~ -25%")
    print("  以 Version B 的權重做實際分配參考")
    if MAX_WEIGHT is None:
        print(f"\n  💡 若集中度過高，可修改檔頭 MAX_WEIGHT = 0.15 (單檔上限 15%)")


if __name__ == '__main__':
    main()
